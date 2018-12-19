from openpyxl import load_workbook
from django.contrib.auth.models import User
from CORE.models import Organization
from HR.models import Employee
from django.contrib.auth.hashers import make_password

def import_Excel(srcFile):
    # wb = load_workbook(r"E:\OneDrive\Work\billjc\区域研发中心\人力资源\在职&离职名单\南京在职&amp;离职名单--2018.5.11.xlsx",data_only=True)
    wb = load_workbook(srcFile, data_only=True )
    sht = wb["人员信息"]
    # colIDX_SRC=["A","B"]
    # colIDX_DES=["A","B"]
    # 花名册Excel表格中，列名称对应的数据库字段属性，key是列名，value是字段属性
    colIDX_Fields_Mapping = {
        "A":"user_id",
        "B":"workNo",
        "C":"workNoExt",
        "D":"userName", #姓名
        "E":"IDNo", #身份证
        "F":"sex", #性别字段
        "G":"level", #性别字段
        "H":"productUnit", #产品线
        "J":"projectName", #项目组
        "L":"role", #员工角色
        "N":"mobilePhone", #手机号
        "O":"majorSkill", #手机号
        "P":"entryDate", # 到岗日期/入职日期
        "Q":"birthDay",# 出生日期
        "R":"graduatedDay", # 毕业时间
        "S":"graduatedSchool", #毕业院校
        "T":"schoolType", #毕业院校类型211/985
        "U":"education", #学历
        "V":"profession",  # 大学专业名称
        "W":"provinceBirth", #籍贯省
        "X":"cityBirth", #籍贯市
        "Y":"maritalStatus",
        "Z":"emailExt", #华为邮箱
        "AA":"email",  #佰钧成邮箱  
        "AB":"workStatus",  #佰钧成邮箱          
        "AC":"address", #居住地址
        "AD":"headPicPath", #员工头像
        }
    usedRowCount = sht.max_row
    # usedRowCount = 10
    #数据从第二行开始读取，第一行是标题
    # for row in range(2,usedRowCount):
    #     p = Employee()
    #     for colName,fieldName  in colIDX_Fields_Mapping.items():
    #         print(colName,fieldName)
    #         #动态设置对象的属性值
    #         p.__setattr__ (fieldName, sht["%s%i" % (colName,row)].value)
    #         # pars[fieldName]=sht["%s%i" % (colName,row)].value
    #     p.save()
    #生成SQL语句
    #数据从第二行开始读取，第一行是标题
    # for row in range(2,usedRowCount):        
    #     fieldSQL = ""
    #     vlsSQL = ""
    #     for colName,fieldName  in colIDX_Fields_Mapping.items():
    #         if fieldSQL !="":
    #             fieldSQL = fieldSQL + ", " + fieldName
    #             vlsSQL   = vlsSQL +(", '%s'")%str( sht["%s%i" % (colName,row)].value)
    #         else:
    #             fieldSQL = fieldName
    #             vlsSQL   = ("'%s'")%str( sht["%s%i" % (colName,row)].value)
            
    #     insSQL = "row=%i=>INSERT INTO HR_employee (%s) values(%s)"%(row,fieldSQL, vlsSQL)
    #     print(insSQL)
    # ------------------------------------------------------------------------------------
    # range(start, stop[, step])
    # start: 计数从 start 开始。默认是从 0 开始。例如range（5）等价于range（0， 5）;
    # stop: 计数到 stop 结束，但不包括 stop。例如：range（0， 5） 是[0, 1, 2, 3, 4]没有5
    # step：步长，默认为1。例如：range（0， 5） 等价于 range(0, 5, 1)
    
    #sheet行号 第一行标题不读取,所以数据值应该减少一次循环
    for row in range(1, usedRowCount ):
        user = User() #创建系统用户对象
        employee = Employee()
        sheetRow = row + 1 #Sheet表的行号和循环下标对应 应该加1
        for colName, fieldName  in colIDX_Fields_Mapping.items():
            v = sht["%s%i" % (colName, sheetRow) ].value
            if fieldName == "user_id":
                user.id = v
            elif fieldName == "workNo":                
                user.username = v
                employee.workNo = v 
                if v == "B-12110":
                    print("workNo==v==%s"%v)
                    user.is_superuser = True
                    user.password= make_password("1qaz@WSX")
                    continue
            elif fieldName == "email":
                user.email = v
                employee.email = v
            elif fieldName == "userName":
                user.first_name = v
                user.last_name = v
                employee.userName = v
                user.password= make_password("123456")

            #读取到项目组列
            elif fieldName == "projectName":
                depart = Organization.objects.get(pk = v)
                employee.depart = depart
            else:
                employee.__setattr__ ( fieldName, sht["%s%i" % (colName, sheetRow)].value )
        user.save()
        employee.user = user
        employee.save()
        msg= ("第%i/%i个，姓名：%s 工号：%s 数据导入完毕。")%( row , (usedRowCount - 1), employee.userName, employee.workNo)
        print(msg)

def initData(*args):
    """初始化数据"""

def main(*args,**kwargs):
    print("===import data===")
    import_Excel(r"demo\testdata.xlsx")


if __name__ == "__main__":
    main()
